import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"sample_data\bca 1st sem.xlsx", data_only=True)
sheet = wb.active

print(f"Sheet: {sheet.title}, Rows: {sheet.max_row}, Cols: {sheet.max_column}")
for r in range(1, sheet.max_row + 1):
    row_vals = [str(sheet.cell(r, c).value).strip() if sheet.cell(r, c).value is not None else "" for c in range(1, sheet.max_column + 1)]
    if any(row_vals):
        while row_vals and row_vals[-1] == "":
            row_vals.pop()
        print(f"R{r:02d}: {row_vals}")
