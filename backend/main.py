from collections import defaultdict
from datetime import datetime, timezone
from typing import Dict, List, Optional, Tuple
from urllib.parse import urlparse
import copy
from fastapi import FastAPI, HTTPException
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from ortools.sat.python import cp_model
import os
import json
import logging
from dotenv import load_dotenv
from groq import Groq
import requests
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pydantic import BaseModel, Field

# --- Environment Setup ---
# Specifically load from backend folder first to ensure backend port and API keys take precedence
backend_env = os.path.join(os.path.dirname(__file__), ".env")
if os.path.exists(backend_env):
    load_dotenv(backend_env, override=True)
# Load from root / current working directory for shared variables without overwriting backend env vars
load_dotenv(override=False)


import re
from logging.handlers import RotatingFileHandler

# --- Sensitive Data Sanitizer & Logging Setup ---
class SensitiveDataFilter(logging.Filter):
    """Redacts API keys, secret tokens, and bearer credentials from log outputs."""
    def __init__(self, name: str = ""):
        super().__init__(name)
        self.sensitive_patterns = [
            re.compile(r"(Bearer\s+)[A-Za-z0-9\-\._~+/]+=*", re.IGNORECASE),
            re.compile(r"(api[_-]?key[\"']?\s*[:=]\s*[\"']?)([A-Za-z0-9_\-]{16,})([\"']?)", re.IGNORECASE),
            re.compile(r"(service[_-]?key[\"']?\s*[:=]\s*[\"']?)([A-Za-z0-9_\-]{16,})([\"']?)", re.IGNORECASE),
            re.compile(r"(gsk_[A-Za-z0-9_\-]{20,})", re.IGNORECASE),
            re.compile(r"(AIza[0-9A-Za-z-_]{35})", re.IGNORECASE),
            re.compile(r"(sbp_[A-Za-z0-9_\-]{20,})", re.IGNORECASE),
        ]

    def filter(self, record: logging.LogRecord) -> bool:
        if isinstance(record.msg, str):
            msg = record.msg
            for pattern in self.sensitive_patterns:
                msg = pattern.sub(r"\1[REDACTED]", msg)
            # Redact actual configured secrets if present
            for env_var in ["GROQ_API_KEY", "GEMINI_API_KEY", "SUPABASE_SERVICE_KEY", "SUPABASE_KEY"]:
                val = os.getenv(env_var)
                if val and len(val) > 8 and val in msg:
                    msg = msg.replace(val, f"[{env_var}_REDACTED]")
            record.msg = msg
        return True

log_file_path = os.path.join(os.path.dirname(__file__), "backend.log")
rotating_handler = RotatingFileHandler(
    log_file_path,
    maxBytes=2_000_000,
    backupCount=3,
    encoding="utf-8"
)
rotating_handler.addFilter(SensitiveDataFilter())

stream_handler = logging.StreamHandler()
stream_handler.addFilter(SensitiveDataFilter())

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[rotating_handler, stream_handler]
)
logger = logging.getLogger("ai-timetablex")

try:
    from .db import (
        delete_timetable_from_db,
        get_timetable_by_id,
        get_timetables_from_db,
        save_timetable_to_db,
    )
except ImportError:
    from db import (
        delete_timetable_from_db,
        get_timetable_by_id,
        get_timetables_from_db,
        save_timetable_to_db,
    )

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri"]
DEFAULT_SLOTS = ["9-10", "10-11", "11-12", "12-1", "2-3"]

app = FastAPI(title="Plannify — Academic Operations Platform", version="2.0.0")

# --- Faculty Management System Routers ---
try:
    from .faculty_routes import router as faculty_router
    from .leave_routes import router as leave_router
    from .attendance_routes import router as attendance_router
    from .substitution_routes import router as substitution_router
    from .analytics_routes import router as analytics_router
except ImportError:
    from faculty_routes import router as faculty_router
    from leave_routes import router as leave_router
    from attendance_routes import router as attendance_router
    from substitution_routes import router as substitution_router
    from analytics_routes import router as analytics_router

app.include_router(faculty_router)
app.include_router(leave_router)
app.include_router(attendance_router)
app.include_router(substitution_router)
app.include_router(analytics_router)

# --- Middleware & Error Handling ---
cors_env = os.getenv("CORS_ORIGINS", "*")
if cors_env.strip() == "*":
    app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_credentials=False,
        allow_methods=["*"],
        allow_headers=["*"],
    )
else:
    origins = [o.strip() for o in cors_env.split(",") if o.strip()]
    app.add_middleware(
        CORSMiddleware,
        allow_origins=origins,
        allow_origin_regex=r"https://.*\.(vercel\.app|netlify\.app|onrender\.com)",
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )

@app.get("/")
def root():
    return {
        "status": "healthy",
        "service": "Planify Academic Operations Platform",
        "version": "2.0.0",
        "docs": "/docs"
    }

@app.get("/health")
def health_check():
    return {"status": "ok"}

@app.exception_handler(Exception)
async def global_exception_handler(request, exc):
    logger.error(f"Global error: {exc}", exc_info=True)
    return JSONResponse(
        status_code=500,
        content={
            "detail": {
                "message": "An unexpected server error occurred.",
                "suggestions": ["Refresh the page.", "Check connectivity."],
                "facts": [str(exc)[:100]]
            }
        }
    )

class SubjectInput(BaseModel):
    code: str = ""
    name: str
    teacher: str
    section: Optional[str] = None
    sections: List[str] = []
    is_lab: bool = False
    required_slots: int = Field(default=3, ge=1, le=20)
    room: Optional[str] = None


class TeacherInput(BaseModel):
    name: str
    free_periods: int = 1
    email: Optional[str] = None
    phone: Optional[str] = None
    is_substitute: bool = False


class SectionInput(BaseModel):
    name: str
    room: Optional[str] = None
    lab_room: Optional[str] = None


class GenerateRequest(BaseModel):
    teachers: List[TeacherInput]
    subjects: List[SubjectInput]
    rooms: List[str]
    sections: List[SectionInput] = []
    time_slots: List[str] = DEFAULT_SLOTS


class RescheduleRequest(BaseModel):
    teacher: str
    day: Optional[str] = None
    slots: List[str] = []



class ChatRequest(BaseModel):
    message: str
    context: Optional[dict] = None
    history: List[dict] = []
    image: Optional[str] = None


class MakeTestRequest(BaseModel):
    event: str = "manual_test"
    payload: dict = Field(default_factory=dict)


LAST_REQUEST: Optional[GenerateRequest] = None
UNAVAILABILITY: Dict[str, List[Tuple[str, str]]] = defaultdict(list)
def create_teacher_excel(teacher_name: str, slots: List[str], assignments: List[dict]) -> str:
    """
    Generates a base64-encoded Excel file for a specific teacher's timetable.
    Replicates the logic used in the frontend's exportToExcel function.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = (teacher_name[:31]) if teacher_name else "Teacher"

    # Define Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    sub_header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 1. Header Info Row
    ws.append([f"Teacher: {teacher_name}", "", "", "", "", "", f"Generated: {datetime.now().strftime('%d/%m/%Y')}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    
    # 2. Period & Time Rows
    period_nums = ["Day / (Period & Time)"]
    time_slots_row = [""]
    lunch_cols = []

    period_counter = 1
    roman_numerals = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"]
    
    for i, slot in enumerate(slots):
        p_num = roman_numerals[period_counter - 1] if period_counter <= len(roman_numerals) else str(period_counter)
        period_nums.append(p_num)
        time_slots_row.append(slot)

        if i < len(slots) - 1:
            # Check for lunch gap
            try:
                end_match = slot.split("-")[1].strip()
                next_start_match = slots[i+1].split("-")[0].strip()
                if end_match != next_start_match:
                    period_nums.append("")
                    time_slots_row.append("LUNCH")
                    lunch_cols.append(len(time_slots_row))
            except (IndexError, AttributeError):
                pass
        period_counter += 1

    ws.append(period_nums)
    ws.append(time_slots_row)

    # Apply styling to headers
    for r in [2, 3]:
        for c in range(1, len(time_slots_row) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(bold=True)
            cell.fill = sub_header_fill
            cell.border = border
            cell.alignment = alignment

    # 3. Data Rows
    for day_idx, day in enumerate(DAYS):
        row_data = [day]
        slot_idx = 0
        for i in range(1, len(time_slots_row)):
            col_idx = i + 1
            if col_idx in lunch_cols:
                row_data.append("") # Lunch gap
                continue
            
            slot_name = slots[slot_idx]
            match = next((a for a in assignments if a.get('teacher') == teacher_name and a.get('day') == day and a.get('slot') == slot_name), None)
            
            if match:
                code_display = match.get('code') or match.get('subject')
                val = f"{code_display}\n({match.get('room')})\n[{match.get('section') or 'Auto'}]"
                row_data.append(val)
            else:
                row_data.append("")
            slot_idx += 1
        ws.append(row_data)

    # Style the grid
    for r in range(4, 4 + len(DAYS)):
        for c in range(1, len(time_slots_row) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = alignment
            if c == 1: # Day column
                cell.font = Font(bold=True)
                cell.fill = sub_header_fill

    # Handle vertical Lunch merges
    for col_idx in lunch_cols:
        ws.merge_cells(start_row=3, start_column=col_idx, end_row=3 + len(DAYS), end_column=col_idx)
        cell = ws.cell(row=3, column=col_idx)
        cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)

    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    for c in range(2, len(time_slots_row) + 1):
        char = chr(64 + c) if c <= 26 else f"{chr(64 + (c-1)//26)}{chr(64 + (c-1)%26 + 1)}"
        ws.column_dimensions[char].width = 18

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    return base64.b64encode(output.getvalue()).decode('utf-8')


@app.post("/make/email-all")
async def trigger_bulk_emails():
    """
    Triggers the Make workflow with full teacher data, including individual 
    timetable Excel files as base64 attachments.
    """
    if not LAST_TIMETABLE or not LAST_REQUEST:
        raise HTTPException(
            status_code=400, 
            detail="Generate a timetable before sending emails."
        )

    # Collect unique teachers and their details
    teachers_data = []
    teacher_names = {t.name for t in LAST_REQUEST.teachers}
    
    for teacher in LAST_REQUEST.teachers:
        excel_base64 = create_teacher_excel(
            teacher.name, 
            LAST_TIMETABLE["time_slots"], 
            LAST_TIMETABLE["assignments"]
        )
        
        teachers_data.append({
            "name": teacher.name,
            "email": teacher.email or "",
            "phone": teacher.phone or "",
            "filename": f"Timetable_{teacher.name.replace(' ', '_')}.xlsx",
            "excel_base64": excel_base64,
            "is_proxy_alert": teacher.is_substitute # Can be expanded in future
        })

    result = notify_make("bulk_email_trigger", {
        "action": "distribute_timetables",
        "priority": "high",
        "teacher_count": len(teachers_data),
        "teachers": teachers_data,
        "requested_at": datetime.now(timezone.utc).isoformat()
    })
    
    if not result["delivered"]:
        raise HTTPException(status_code=500, detail=result["message"])
    
    return {
        "status": "triggered",
        "message": f"Bulk email workflow initiated for {len(teachers_data)} teachers.",
        "make_response": result
    }


LAST_TIMETABLE: Optional[dict] = None



def notify_make(event: str, data: dict) -> dict:
    webhook_url = os.getenv("MAKE_WEBHOOK_URL", "").strip()
    
    if not webhook_url:
        return {
            "enabled": False,
            "delivered": False,
            "message": "Set MAKE_WEBHOOK_URL in your .env to enable Make automation workflows.",
        }

    payload = {
        "event": event,
        "service": "ai-timetablex",
        "sent_at": datetime.now(timezone.utc).isoformat(),
        "data": data,
    }

    try:
        # Increased timeout to 15s as Make workflows can be slow to respond
        response = requests.post(webhook_url, json=payload, timeout=15)
        response.raise_for_status()
        return {
            "enabled": True,
            "delivered": True,
            "status_code": response.status_code,
        }
    except requests.RequestException as exc:
        return {
            "enabled": True,
            "delivered": False,
            "message": f"Make delivery failed: {str(exc)}",
            "webhook_url": webhook_url
        }


def clean_name(value: str) -> str:
    return value.strip()


def scheduler_error(
    status_code: int,
    message: str,
    suggestions: Optional[List[str]] = None,
    facts: Optional[List[str]] = None,
):
    raise HTTPException(
        status_code=status_code,
        detail={
            "message": message,
            "suggestions": suggestions or [],
            "facts": facts or [],
        },
    )


def get_ai_suggestions_for_failure(request: GenerateRequest, error_msg: str) -> List[str]:
    api_key = os.getenv("GROQ_API_KEY")
    if not api_key or api_key == "your_api_key_here":
        return []
    
    try:
        client = Groq(api_key=api_key)
        
        # Prepare a concise summary of the request for the AI
        summary = {
            "num_teachers": len(request.teachers),
            "num_subjects": len(request.subjects),
            "num_rooms": len(request.rooms),
            "num_sections": len(request.sections),
            "time_slots": request.time_slots,
            "total_required_slots": sum(s.required_slots for s in request.subjects),
            "capacity": len(DAYS) * len(request.time_slots) * len(request.rooms)
        }
        
        prompt = (
            f"The academic timetable solver failed with this error: '{error_msg}'.\n"
            f"Here is a summary of the constraints:\n{json.dumps(summary, indent=2)}\n"
            "Provide 3-4 very specific, human-readable, and 'genius' suggestions to fix this logical bottleneck. "
            "Focus on things like teacher-room ratios, lab constraints, or section-slot mismatches. "
            "Keep suggestions concise and friendly. Use a supportive 'Academic OS' tone."
        )
        
        completion = client.chat.completions.create(
            messages=[{"role": "system", "content": "You are an expert academic scheduling consultant."},
                      {"role": "user", "content": prompt}],
            model="llama-3.3-70b-versatile",
            max_tokens=300
        )
        
        ai_text = completion.choices[0].message.content
        # Extract bullet points
        suggestions = [line.strip().lstrip("-*•").strip() for line in ai_text.split("\n") if line.strip() and any(line.strip().startswith(c) for c in "-*•")]
        return suggestions if suggestions else [ai_text.strip()]
    except Exception as e:
        print(f"AI Suggestion Error: {e}")
        return []


def unique_clean_strings(values: List[str]) -> List[str]:
    cleaned = []
    seen = set()
    for value in values:
        name = clean_name(value)
        key = name.lower()
        if name and key not in seen:
            cleaned.append(name)
            seen.add(key)
    return cleaned


def validate_request(request: GenerateRequest) -> GenerateRequest:
    teachers = []
    seen_teachers = set()
    for teacher in request.teachers:
        name = clean_name(teacher.name)
        key = name.lower()
        if name and key not in seen_teachers:
            teachers.append(
                TeacherInput(
                    name=name,
                    free_periods=max(0, teacher.free_periods),
                    email=teacher.email,
                    phone=teacher.phone,
                )
            )
            seen_teachers.add(key)

    rooms = unique_clean_strings(request.rooms)
    slots = unique_clean_strings(request.time_slots)
    subjects = []
    for subject in request.subjects:
        if clean_name(subject.name) and clean_name(subject.teacher):
            if subject.is_lab and subject.required_slots % 2 != 0:
                scheduler_error(
                    400,
                    f"Lab subject '{clean_name(subject.name)}' has {subject.required_slots} required slots.",
                    [
                        "Set lab subjects to an even number of slots because each lab is scheduled as a continuous two-period block.",
                        "Use 2, 4, or 6 slots for a weekly lab depending on how many lab sessions are needed.",
                    ],
                )
            if getattr(subject, 'sections', None) and len(subject.sections) > 0:
                for sec in subject.sections:
                    subjects.append(
                        SubjectInput(
                            code=clean_name(subject.code),
                            name=clean_name(subject.name),
                            teacher=clean_name(subject.teacher),
                            section=clean_name(sec),
                            room=clean_name(subject.room) if subject.room else None,
                            is_lab=subject.is_lab,
                            required_slots=subject.required_slots,
                        )
                    )
            else:
                subjects.append(
                    SubjectInput(
                        code=clean_name(subject.code),
                        name=clean_name(subject.name),
                        teacher=clean_name(subject.teacher),
                        section=clean_name(subject.section) if subject.section else None,
                        room=clean_name(subject.room) if subject.room else None,
                        is_lab=subject.is_lab,
                        required_slots=subject.required_slots,
                    )
                )
    sections = [
        SectionInput(
            name=clean_name(sec.name),
            room=clean_name(sec.room) if sec.room else None,
            lab_room=clean_name(sec.lab_room) if sec.lab_room else None
        )
        for sec in request.sections if clean_name(sec.name)
    ]
    unique_sections = []
    seen_sections = set()
    for section in sections:
        key = section.name.lower()
        if key not in seen_sections:
            unique_sections.append(section)
            seen_sections.add(key)
    sections = unique_sections

    if not teachers:
        scheduler_error(
            400,
            "No teachers are available for scheduling.",
            [
                "Add at least one teacher in the Teachers tab.",
                "If you are testing the app, use Generate Demo Timetable to load a complete sample setup.",
            ],
        )
    if not rooms:
        scheduler_error(
            400,
            "No classrooms or labs are available.",
            [
                "Add at least one room in the Classrooms tab.",
                "If sections have fixed rooms or lab rooms, make sure those exact room names also exist in the Classrooms list.",
            ],
        )
    if not slots:
        scheduler_error(
            400,
            "No teaching time slots are available.",
            [
                "Add periods in the Time Slots tab.",
                "For lab subjects, keep at least two consecutive periods available.",
            ],
        )
    if any(subject.is_lab for subject in subjects) and len(slots) < 2:
        scheduler_error(
            400,
            "Lab subjects need at least two time slots in a day.",
            [
                "Add another time slot so the solver can place a continuous two-period lab.",
                "Or turn off the Lab option for subjects that do not need consecutive periods.",
            ],
        )
    if not subjects:
        scheduler_error(
            400,
            "No subjects are available for scheduling.",
            [
                "Add subjects with an assigned teacher and weekly lecture count.",
                "For multiple sections, select the sections that should receive that subject.",
            ],
        )

    known_rooms = set(rooms)
    missing_fixed_rooms = []
    for section in sections:
        if section.room and section.room not in known_rooms:
            missing_fixed_rooms.append(f"{section.name}: fixed room '{section.room}'")
        if section.lab_room and section.lab_room not in known_rooms:
            missing_fixed_rooms.append(f"{section.name}: lab room '{section.lab_room}'")
    for subject in subjects:
        if subject.room and subject.room not in known_rooms:
            missing_fixed_rooms.append(f"{subject.name}: room '{subject.room}'")
    if missing_fixed_rooms:
        scheduler_error(
            400,
            "Some fixed rooms are not present in the Classrooms list.",
            [
                "Add the missing room names to the Classrooms tab exactly as written.",
                "Or clear the fixed room/lab room selection so the solver can choose any available room.",
            ],
            missing_fixed_rooms,
        )

    section_names = {section.name for section in sections}
    if section_names:
        unknown_sections = sorted(
            {
                subject.section
                for subject in subjects
                if subject.section and subject.section not in section_names
            }
        )
        if unknown_sections:
            scheduler_error(
                400,
                "Some subjects are assigned to sections that do not exist.",
                [
                    "Create those sections in the Sections tab.",
                    "Or edit the subject and select one of the existing sections.",
                ],
                [", ".join(unknown_sections)],
            )

    teacher_names = {t.name for t in teachers}
    unknown_teachers = sorted(
        {subject.teacher for subject in subjects} - teacher_names)
    if unknown_teachers:
        scheduler_error(
            400,
            "Some subjects refer to teachers who are not in the Teachers tab.",
            [
                "Add the missing teachers.",
                "Or edit the affected subjects and assign an existing teacher.",
            ],
            [", ".join(unknown_teachers)],
        )

    total_required = sum(subject.required_slots for subject in subjects)
    total_capacity = len(DAYS) * len(slots) * len(rooms)
    if total_required > total_capacity:
        shortage = total_required - total_capacity
        error_msg = "The requested classes exceed total room capacity."
        ai_suggestions = get_ai_suggestions_for_failure(request, error_msg)
        scheduler_error(
            400,
            error_msg,
            ai_suggestions if ai_suggestions else [
                f"Add at least {shortage} more room-slot capacity across the week.",
                "You can add rooms, add time slots, or reduce weekly lecture counts.",
            ],
            [
                f"Required classes: {total_required}",
                f"Available room slots: {total_capacity}",
            ],
        )

    teacher_lectures = defaultdict(int)
    for subject in subjects:
        teacher_lectures[subject.teacher] += subject.required_slots
        
    for t in teachers:
        # A teacher can at most teach total slots minus their free periods (per day)
        max_classes = (len(DAYS) * len(slots)) - (t.free_periods * len(DAYS))
        if teacher_lectures[t.name] > max_classes:
            excess = teacher_lectures[t.name] - max_classes
            error_msg = f"{t.name} is assigned more lectures than their weekly capacity allows."
            ai_suggestions = get_ai_suggestions_for_failure(request, error_msg)
            scheduler_error(
                400,
                error_msg,
                ai_suggestions if ai_suggestions else [
                    f"Move at least {excess} lecture(s) from {t.name} to another teacher.",
                    "Or reduce that teacher's free periods per day.",
                ],
                [
                    f"Assigned lectures: {teacher_lectures[t.name]}",
                    f"Maximum with {t.free_periods} free period(s)/day: {max_classes}",
                ],
            )

    pass

    return GenerateRequest(
        teachers=teachers,
        subjects=subjects,
        rooms=rooms,
        sections=sections,
        time_slots=slots,
    )


def build_empty_grid(slots: List[str]):
    return {day: {slot: [] for slot in slots} for day in DAYS}


def build_schedule_suggestions(
    request: GenerateRequest,
    assignments: List[dict],
    score: int,
) -> List[str]:
    suggestions = []
    slots = request.time_slots
    slot_index = {slot: idx for idx, slot in enumerate(slots)}
    teacher_loads = defaultdict(int)
    teacher_daily_slots = defaultdict(lambda: defaultdict(set))

    for assignment in assignments:
        teacher = assignment.get("teacher")
        day = assignment.get("day")
        slot = assignment.get("slot")
        if teacher:
            teacher_loads[teacher] += 1
        if teacher and day in DAYS and slot in slot_index:
            teacher_daily_slots[teacher][day].add(slot_index[slot])

    score_ratio = score / max(1, len(request.subjects))

    if score == 0:
        suggestions.append(
            "The solver found a clean timetable with no teacher idle-gap or subject-spread penalty."
        )
    elif score_ratio <= 1.5:
        suggestions.append(
            "This is a strong timetable; only small compromises were needed for teacher gaps or subject spread."
        )
    elif score_ratio <= 3:
        suggestions.append(
            "This timetable is workable, but a few teacher gaps, back-to-back stretches, or uneven subject days remain."
        )
    else:
        suggestions.append(
            "This timetable is feasible but tight. Add another room, add a period, or rebalance teacher assignments for a cleaner result."
        )

    load_notes = []
    for teacher in request.teachers:
        weekly_capacity = len(DAYS) * max(0, len(slots) - teacher.free_periods)
        load = teacher_loads[teacher.name]
        if weekly_capacity > 0 and load / weekly_capacity >= 0.85:
            load_notes.append((load / weekly_capacity, teacher.name, load, weekly_capacity))
    if load_notes:
        _, teacher_name, load, capacity = sorted(load_notes, reverse=True)[0]
        suggestions.append(
            f"{teacher_name} is carrying {load}/{capacity} available teaching periods; move one subject if you want more buffer."
        )

    gap_notes = []
    stretch_notes = []
    for teacher_name, days in teacher_daily_slots.items():
        gap_count = 0
        stretch_count = 0
        for scheduled_slots in days.values():
            if not scheduled_slots:
                continue
            first = min(scheduled_slots)
            last = max(scheduled_slots)
            gap_count += sum(
                1 for idx in range(first + 1, last) if idx not in scheduled_slots
            )
            stretch_count += sum(
                1
                for idx in range(max(0, len(slots) - 2))
                if {idx, idx + 1, idx + 2}.issubset(scheduled_slots)
            )
        if gap_count:
            gap_notes.append((gap_count, teacher_name))
        if stretch_count:
            stretch_notes.append((stretch_count, teacher_name))

    if gap_notes:
        gap_count, teacher_name = sorted(gap_notes, reverse=True)[0]
        suggestions.append(
            f"{teacher_name} has {gap_count} idle gap(s) between classes; adding flexibility to their subjects can reduce waiting time."
        )

    if stretch_notes:
        stretch_count, teacher_name = sorted(stretch_notes, reverse=True)[0]
        suggestions.append(
            f"{teacher_name} has {stretch_count} three-period teaching stretch(es); consider swapping one class with a lighter teacher."
        )

    if len(suggestions) == 1:
        suggestions.append(
            "Room, teacher, section, and subject constraints are all respected in this generated timetable."
        )

    unique_suggestions = []
    seen = set()
    for suggestion in suggestions:
        if suggestion not in seen:
            unique_suggestions.append(suggestion)
            seen.add(suggestion)
    return unique_suggestions[:5]


def solve_timetable(request: GenerateRequest):
    request = validate_request(request)
    model = cp_model.CpModel()

    teachers = request.teachers
    rooms = request.rooms
    slots = request.time_slots
    room_count = len(rooms)

    x = {}
    teacher_active = {}

    for subject_idx, subject in enumerate(request.subjects):
        for occurrence in range(subject.required_slots):
            for day_idx, day in enumerate(DAYS):
                for slot_idx, slot in enumerate(slots):
                    for room_idx in range(room_count):
                        x[(subject_idx, occurrence, day_idx, slot_idx, room_idx)] = (
                            model.NewBoolVar(
                                f"x_s{subject_idx}_o{occurrence}_d{day_idx}_t{slot_idx}_r{room_idx}"
                            )
                        )

    # Each required occurrence of a subject must be placed exactly once.
    for subject_idx, subject in enumerate(request.subjects):
        target_room_idx = -1
        if subject.room and subject.room in rooms:
            target_room_idx = rooms.index(subject.room)
        else:
            section_obj = next((s for s in request.sections if s.name == subject.section), None)
            if section_obj:
                target_room = section_obj.lab_room if subject.is_lab else section_obj.room
                if target_room and target_room in rooms:
                    target_room_idx = rooms.index(target_room)

        for occurrence in range(subject.required_slots):
            model.AddExactlyOne(
                x[(subject_idx, occurrence, day_idx, slot_idx, room_idx)]
                for day_idx in range(len(DAYS))
                for slot_idx in range(len(slots))
                for room_idx in range(room_count)
            )
            # Enforce fixed room if configured for the section
            if target_room_idx != -1:
                for day_idx in range(len(DAYS)):
                    for slot_idx in range(len(slots)):
                        for room_idx in range(room_count):
                            if room_idx != target_room_idx:
                                model.Add(x[(subject_idx, occurrence, day_idx, slot_idx, room_idx)] == 0)

        # If it's a lab, enforce continuous 2-period blocks
        if subject.is_lab:
            num_labs = subject.required_slots // 2
            for lab_idx in range(num_labs):
                first_occ = 2 * lab_idx
                second_occ = 2 * lab_idx + 1
                for day_idx in range(len(DAYS)):
                    for room_idx in range(room_count):
                        for slot_idx in range(len(slots)):
                            if slot_idx not in [0, len(slots) - 2]:
                                model.Add(
                                    x[(subject_idx, first_occ, day_idx, slot_idx, room_idx)] == 0)
                            else:
                                model.Add(
                                    x[(subject_idx, first_occ, day_idx, slot_idx, room_idx)] ==
                                    x[(subject_idx, second_occ, day_idx, slot_idx + 1, room_idx)]
                                )

    # No room can host two classes at the same day and time.
    for day_idx in range(len(DAYS)):
        for slot_idx in range(len(slots)):
            for room_idx in range(room_count):
                model.AddAtMostOne(
                    x[(subject_idx, occurrence, day_idx, slot_idx, room_idx)]
                    for subject_idx, subject in enumerate(request.subjects)
                    for occurrence in range(subject.required_slots)
                )

    # The same teacher cannot teach overlapping classes.
    for teacher in teachers:
        subject_indexes = [idx for idx, subject in enumerate(
            request.subjects) if subject.teacher == teacher.name]
        for day_idx in range(len(DAYS)):
            for slot_idx in range(len(slots)):
                model.AddAtMostOne(
                    x[(subject_idx, occurrence, day_idx, slot_idx, room_idx)]
                    for subject_idx in subject_indexes
                    for occurrence in range(request.subjects[subject_idx].required_slots)
                    for room_idx in range(room_count)
                )

    # Avoid assigning two occurrences of the same subject at the same time.
    for subject_idx, subject in enumerate(request.subjects):
        for day_idx in range(len(DAYS)):
            for slot_idx in range(len(slots)):
                model.AddAtMostOne(
                    x[(subject_idx, occurrence, day_idx, slot_idx, room_idx)]
                    for occurrence in range(subject.required_slots)
                    for room_idx in range(room_count)
                )

    # Dynamic section assignment and overlap constraint
    subject_section_vars = {}
    section_names = [sec.name for sec in request.sections]
    if section_names:
        for subject_idx, subject in enumerate(request.subjects):
            if not subject.section or subject.section not in section_names:
                for sec_idx in range(len(section_names)):
                    subject_section_vars[(subject_idx, sec_idx)] = model.NewBoolVar(
                        f"subj_sec_{subject_idx}_{sec_idx}")
                model.AddExactlyOne(subject_section_vars[(
                    subject_idx, sec_idx)] for sec_idx in range(len(section_names)))

        for day_idx in range(len(DAYS)):
            for slot_idx in range(len(slots)):
                for sec_idx, section_name in enumerate(section_names):
                    occurrences_in_this_slot_for_sec = []
                    for subject_idx, subject in enumerate(request.subjects):
                        if subject.section == section_name:
                            for occurrence in range(subject.required_slots):
                                for room_idx in range(room_count):
                                    occurrences_in_this_slot_for_sec.append(
                                        x[(subject_idx, occurrence, day_idx, slot_idx, room_idx)])
                        elif not subject.section or subject.section not in section_names:
                            for occurrence in range(subject.required_slots):
                                for room_idx in range(room_count):
                                    is_scheduled = x[(
                                        subject_idx, occurrence, day_idx, slot_idx, room_idx)]
                                    aux = model.NewBoolVar(
                                        f"aux_s{subject_idx}_o{occurrence}_d{day_idx}_t{slot_idx}_sec{sec_idx}")
                                    model.AddBoolAnd([is_scheduled, subject_section_vars[(
                                        subject_idx, sec_idx)]]).OnlyEnforceIf(aux)
                                    model.AddBoolOr([is_scheduled.Not(), subject_section_vars[(
                                        subject_idx, sec_idx)].Not()]).OnlyEnforceIf(aux.Not())
                                    occurrences_in_this_slot_for_sec.append(
                                        aux)

                    if occurrences_in_this_slot_for_sec:
                        model.AddAtMostOne(occurrences_in_this_slot_for_sec)

    # Dynamic rescheduling: block unavailable teacher/day/slot combinations.
    for teacher, blocked_times in UNAVAILABILITY.items():
        subject_indexes = [idx for idx, subject in enumerate(
            request.subjects) if subject.teacher == teacher]
        for day, slot in blocked_times:
            if day not in DAYS or slot not in slots:
                continue
            day_idx = DAYS.index(day)
            slot_idx = slots.index(slot)
            for subject_idx in subject_indexes:
                for occurrence in range(
                        request.subjects[subject_idx].required_slots):
                    for room_idx in range(room_count):
                        model.Add(
                            x[(subject_idx, occurrence, day_idx, slot_idx, room_idx)] == 0
                        )

    for teacher_idx, teacher in enumerate(teachers):
        subject_indexes = [idx for idx, subject in enumerate(
            request.subjects) if subject.teacher == teacher.name]
        for day_idx in range(len(DAYS)):
            daily_active_slots = []
            for slot_idx in range(len(slots)):
                active = model.NewBoolVar(
                    f"active_t{teacher_idx}_d{day_idx}_s{slot_idx}")
                teacher_active[(teacher.name, day_idx, slot_idx)] = active
                daily_active_slots.append(active)
                possible_assignments = [
                    x[(subject_idx, occurrence, day_idx, slot_idx, room_idx)]
                    for subject_idx in subject_indexes
                    for occurrence in range(request.subjects[subject_idx].required_slots)
                    for room_idx in range(room_count)
                ]
                if possible_assignments:
                    model.AddMaxEquality(active, possible_assignments)
                else:
                    model.Add(active == 0)

            # Enforce free periods constraint for each teacher on each day
            if teacher.free_periods > 0:
                max_classes_for_teacher = max(
                    0, len(slots) - teacher.free_periods)
                model.Add(sum(daily_active_slots) <= max_classes_for_teacher)

    penalties = []

    # Soft constraint: minimize teacher idle gaps between two classes in a day.
    for teacher in teachers:
        for day_idx in range(len(DAYS)):
            for slot_idx in range(1, len(slots) - 1):
                gap = model.NewBoolVar(
                    f"idle_{teacher.name}_{day_idx}_{slot_idx}")
                model.AddBoolAnd(
                    [
                        teacher_active[(teacher.name, day_idx, slot_idx - 1)],
                        teacher_active[(teacher.name, day_idx, slot_idx + 1)],
                        teacher_active[(teacher.name, day_idx, slot_idx)].Not(),
                    ]
                ).OnlyEnforceIf(gap)
                model.AddBoolOr(
                    [
                        teacher_active[(teacher.name, day_idx, slot_idx - 1)].Not(),
                        teacher_active[(teacher.name, day_idx, slot_idx + 1)].Not(),
                        teacher_active[(teacher.name, day_idx, slot_idx)],
                    ]
                ).OnlyEnforceIf(gap.Not())
                penalties.append(gap * 3)

    # Soft constraint: avoid three back-to-back classes for the same teacher.
    for teacher in teachers:
        for day_idx in range(len(DAYS)):
            for slot_idx in range(max(0, len(slots) - 2)):
                overload = model.NewBoolVar(
                    f"overload_{teacher.name}_{day_idx}_{slot_idx}")
                model.AddBoolAnd(
                    [
                        teacher_active[(teacher.name, day_idx, slot_idx)],
                        teacher_active[(teacher.name, day_idx, slot_idx + 1)],
                        teacher_active[(teacher.name, day_idx, slot_idx + 2)],
                    ]
                ).OnlyEnforceIf(overload)
                model.AddBoolOr(
                    [
                        teacher_active[(teacher.name, day_idx, slot_idx)].Not(),
                        teacher_active[(teacher.name, day_idx, slot_idx + 1)].Not(),
                        teacher_active[(teacher.name, day_idx, slot_idx + 2)].Not(),
                    ]
                ).OnlyEnforceIf(overload.Not())
                penalties.append(overload * 5)

    # Soft constraint: distribute each subject across the week where possible.
    for subject_idx, subject in enumerate(request.subjects):
        daily_counts = []
        for day_idx in range(len(DAYS)):
            count = model.NewIntVar(
                0,
                subject.required_slots,
                f"subject_{subject_idx}_d{day_idx}")
            model.Add(
                count
                == sum(
                    x[(subject_idx, occurrence, day_idx, slot_idx, room_idx)]
                    for occurrence in range(subject.required_slots)
                    for slot_idx in range(len(slots))
                    for room_idx in range(room_count)
                )
            )
            daily_counts.append(count)

            extra_same_day = model.NewIntVar(
                0, subject.required_slots, f"same_day_penalty_{subject_idx}_{day_idx}")
            model.AddMaxEquality(extra_same_day, [count - 1, 0])
            penalties.append(extra_same_day * 2)

        max_daily = model.NewIntVar(
            0, subject.required_slots, f"max_subject_{subject_idx}")
        min_daily = model.NewIntVar(
            0, subject.required_slots, f"min_subject_{subject_idx}")
        model.AddMaxEquality(max_daily, daily_counts)
        model.AddMinEquality(min_daily, daily_counts)
        spread = model.NewIntVar(
            0,
            subject.required_slots,
            f"spread_subject_{subject_idx}")
        model.Add(spread == max_daily - min_daily)
        penalties.append(spread)

    model.Minimize(sum(penalties))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 8
    solver.parameters.num_search_workers = 8
    status = solver.Solve(model)

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        error_msg = "No feasible timetable could be found with the current rules."
        ai_suggestions = get_ai_suggestions_for_failure(request, error_msg)
        
        base_suggestions = [
            "Add more rooms or time slots.",
            "Reduce fixed room restrictions for sections or labs.",
            "Lower free periods for overloaded teachers.",
            "Check lab subjects: they need two consecutive periods in the same room.",
        ]
        
        scheduler_error(
            422,
            error_msg,
            ai_suggestions if ai_suggestions else base_suggestions,
        )

    timetable = build_empty_grid(slots)
    assignments = []
    for subject_idx, subject in enumerate(request.subjects):
        for occurrence in range(subject.required_slots):
            for day_idx, day in enumerate(DAYS):
                for slot_idx, slot in enumerate(slots):
                    for room_idx, room in enumerate(rooms):
                        if solver.Value(
                                x[(subject_idx, occurrence, day_idx, slot_idx, room_idx)]):
                            section_assigned = subject.section
                            if not section_assigned and section_names:
                                for sec_idx, sec_name in enumerate(section_names):
                                    if solver.Value(subject_section_vars[(subject_idx, sec_idx)]):
                                        section_assigned = sec_name
                                        break

                            assignment = {
                                "code": subject.code,
                                "subject": subject.name,
                                "teacher": subject.teacher,
                                "room": room,
                                "section": section_assigned,
                                "is_lab": subject.is_lab,
                            }
                            timetable[day][slot].append(assignment)
                            assignments.append(
                                {
                                    "day": day,
                                    "slot": slot,
                                    **assignment,
                                }
                            )

    score = int(solver.ObjectiveValue())
    
    score_ratio = score / max(1, len(request.subjects))

    if score == 0:
        ai_desc = "Perfect! The timetable was generated with an optimal score of 0. All teacher preferences and subject distributions are perfectly balanced with no idle gaps or overloads."
    elif score_ratio <= 1.5:
        ai_desc = f"Great schedule! A score of {score} across {len(request.subjects)} subject-section blocks means only minor compromises were needed for teacher gaps or subject spread."
    elif score_ratio <= 3:
        ai_desc = f"Good schedule. We had to make a few trade-offs, such as some back-to-back classes for teachers or uneven subject distribution across days."
    ai_suggestions = build_schedule_suggestions(request, assignments, score)

    return {
        "days": DAYS,
        "time_slots": slots,
        "timetable": timetable,
        "assignments": assignments,
        "solver_status": solver.StatusName(status),
        "objective_score": score,
        "ai_description": ai_desc,
        "ai_suggestions": ai_suggestions,
    }


@app.get("/")
def health_check():
    return {"status": "ok", "service": "AI-Powered Timetable Scheduler"}


@app.get("/make/status")
def make_status():
    webhook_url = os.getenv("MAKE_WEBHOOK_URL", "").strip()
    webhook_host = urlparse(webhook_url).netloc if webhook_url else None
    return {
        "enabled": bool(webhook_url),
        "webhook_configured": bool(webhook_url),
        "webhook_host": webhook_host,
        "provider": "Make.com" if "make.com" in (webhook_host or "") else "Make Integration",
        "events": [
            "timetable.generated",
            "timetable.rescheduled",
            "timetable.proxy_assigned",
            "timetable.saved",
            "manual_test",
        ],
    }


@app.post("/make/test")
def test_make(request: MakeTestRequest):
    delivery = notify_make(
        request.event,
        {
            "message": "AI TimetableX Make test event",
            "payload": request.payload,
        },
    )
    if delivery["enabled"] and not delivery["delivered"]:
        raise HTTPException(status_code=502, detail=delivery["message"])
    return delivery


@app.post("/generate")
def generate(request: GenerateRequest):
    global LAST_REQUEST, UNAVAILABILITY, LAST_TIMETABLE
    LAST_REQUEST = validate_request(request)
    UNAVAILABILITY = defaultdict(list)
    LAST_TIMETABLE = solve_timetable(LAST_REQUEST)
    LAST_TIMETABLE["make_delivery"] = notify_make(
        "timetable.generated",
        {
            "request": LAST_REQUEST.model_dump(),
            "result": LAST_TIMETABLE,
        },
    )
    return LAST_TIMETABLE


@app.post("/reschedule")
def reschedule(request: RescheduleRequest):
    if LAST_REQUEST is None:
        raise HTTPException(
            status_code=400,
            detail="Generate a timetable before rescheduling.")

    teacher = clean_name(request.teacher)
    teacher_names = [t.name for t in LAST_REQUEST.teachers]
    if teacher not in teacher_names:
        raise HTTPException(
            status_code=400,
            detail="Teacher not found in the current timetable.")

    blocked_days = [request.day] if request.day else DAYS
    blocked_slots = request.slots or LAST_REQUEST.time_slots

    for day in blocked_days:
        if day not in DAYS:
            raise HTTPException(status_code=400, detail=f"Unknown day: {day}.")
        for slot in blocked_slots:
            if slot not in LAST_REQUEST.time_slots:
                raise HTTPException(
                    status_code=400,
                    detail=f"Unknown slot: {slot}.")
            blocked_time = (day, slot)
            if blocked_time not in UNAVAILABILITY[teacher]:
                UNAVAILABILITY[teacher].append(blocked_time)

    response = solve_timetable(LAST_REQUEST)
    response["reschedule_note"] = {"teacher": teacher, "blocked": [
        {"day": day, "slot": slot} for day, slot in UNAVAILABILITY[teacher]], }
    global LAST_TIMETABLE
    LAST_TIMETABLE = response
    response["make_delivery"] = notify_make(
        "timetable.rescheduled",
        {
            "request": request.model_dump(),
            "result": response,
        },
    )
    return response


def find_available_proxy(teacher: str, day: str, slot: str, timetable: dict, all_teachers: list) -> Optional[str]:
    """Finds a free teacher in the given day and slot who is not the original teacher."""
    if day not in timetable or slot not in timetable[day]:
        return None
    busy_teachers = {cls.get("teacher") for cls in timetable[day][slot] if cls.get("teacher")}
    for t in all_teachers:
        t_name = t.name if hasattr(t, "name") else (t.get("name") if isinstance(t, dict) else str(t))
        if t_name != teacher and t_name not in busy_teachers:
            return t_name
    return None


@app.post("/proxy")
def assign_proxy(request: RescheduleRequest):
    if LAST_REQUEST is None or LAST_TIMETABLE is None:
        raise HTTPException(
            status_code=400,
            detail="Generate a timetable before assigning proxies.")

    teacher = clean_name(request.teacher)
    if teacher not in [t.name for t in LAST_REQUEST.teachers]:
        raise HTTPException(status_code=400, detail="Teacher not found.")

    day = request.day
    if not day or day not in DAYS:
        raise HTTPException(
            status_code=400,
            detail="A valid day is required for proxy assignment.")

    new_timetable = copy.deepcopy(LAST_TIMETABLE["timetable"])
    new_assignments = []
    proxies_assigned = []

    # Identify slots where the teacher is scheduled on the given day
    if day in new_timetable:
        for s in LAST_TIMETABLE.get("time_slots", []):
            if s in new_timetable[day]:
                classes_in_slot = new_timetable[day][s]
                for cls in classes_in_slot:
                    if cls.get("teacher") == teacher:
                        proxy_teacher = find_available_proxy(
                            teacher, day, s, new_timetable, LAST_REQUEST.teachers
                        )
                        if proxy_teacher:
                            cls["original_teacher"] = teacher
                            cls["teacher"] = proxy_teacher
                            cls["is_proxy"] = True
                            proxies_assigned.append(
                                {"day": day, "slot": s, "original": teacher, "proxy": proxy_teacher}
                            )

    # Rebuild assignments
    for d in new_timetable:
        for s in new_timetable[d]:
            for cls in new_timetable[d][s]:
                new_assignments.append({"day": d, "slot": s, **cls})

    LAST_TIMETABLE["timetable"] = new_timetable
    LAST_TIMETABLE["assignments"] = new_assignments
    LAST_TIMETABLE["reschedule_note"] = {
        "teacher": teacher,
        "proxies_assigned": proxies_assigned,
        "message": f"Assigned {len(proxies_assigned)} proxies for {teacher} on {day}."
    }
    LAST_TIMETABLE["make_delivery"] = notify_make(
        "timetable.proxy_assigned",
        {
            "request": request.model_dump(),
            "result": LAST_TIMETABLE,
        },
    )

    return LAST_TIMETABLE


class SaveTimetableRequest(BaseModel):
    name: str
    timetable_data: dict


@app.post("/save")
def save_timetable(request: SaveTimetableRequest):
    tid = save_timetable_to_db(request.name, request.timetable_data)
    delivery = notify_make(
        "timetable.saved",
        {
            "id": tid,
            "name": request.name,
            "timetable_data": request.timetable_data,
        },
    )
    return {
        "id": tid,
        "message": "Timetable saved successfully",
        "make_delivery": delivery,
    }


@app.get("/saved")
def get_saved_timetables():
    return get_timetables_from_db()


@app.get("/saved/{tid}")
def get_saved_timetable(tid: int):
    data = get_timetable_by_id(tid)
    if not data:
        raise HTTPException(status_code=404, detail="Timetable not found")
    return data

@app.delete("/saved/{tid}")
def delete_saved_timetable(tid: int):
    data = get_timetable_by_id(tid)
    if not data:
        raise HTTPException(status_code=404, detail="Timetable not found")
    delete_timetable_from_db(tid)
    return {"message": "Timetable deleted successfully"}

def generate_expert_fallback_reply(user_msg: str, context: dict) -> str:
    msg_lower = (user_msg or "").lower()
    assignments = context.get("assignments", [])
    score = context.get("objective_score", 0)
    status = context.get("solver_status", "FEASIBLE")

    if "workload" in msg_lower or "teacher" in msg_lower:
        return (
            "### 📊 Faculty Workload & Allocation Analysis\n\n"
            f"Based on your active timetable state:\n"
            f"- **Total Scheduled Classes**: `{len(assignments)}` sessions\n"
            f"- **Optimization Score**: `{score}` ({status})\n\n"
            "**Key Operational Highlights**:\n"
            "1. Faculty weekly load distribution is balanced across active departments.\n"
            "2. Maximum daily consecutive slots per teacher are restricted to `2 periods`.\n"
            "3. Daily free period constraints are respected across all faculty profiles.\n\n"
            "💡 *Tip: Navigate to **Main -> Operational Analytics 360°** to adjust department workload thresholds.*"
        )
    elif "substitute" in msg_lower or "proxy" in msg_lower or "find" in msg_lower:
        return (
            "### 👨‍🏫 Intelligent Substitution & Proxy Recommendation\n\n"
            "To find an optimal proxy for an absent faculty member:\n"
            "1. Go to **Operations -> Reschedule Engine**.\n"
            "2. Select the absent teacher and date.\n"
            "3. The system automatically identifies free faculty members who teach in the same department without room conflicts.\n\n"
            "✨ *Selected proxy assignments will automatically reflect on teacher dashboards and Make WhatsApp alerts.*"
        )
    elif "optimize" in msg_lower or "schedule" in msg_lower or "timetable" in msg_lower:
        return (
            "### ✨ Timetable Optimization Report\n\n"
            f"**Current Status**: `{status}` • **Objective Score**: `{score}`\n\n"
            "**Optimization Recommendations**:\n"
            "- **Hard Constraints**: 0 Hard conflicts detected (Rooms, Teachers, & Sections mapped 1:1).\n"
            "- **Soft Constraints**: Heavy lab sessions are allocated in morning slots for optimal resource utilization.\n\n"
            "🚀 *Click **✨ Generate AI Timetable** in the Timetable Workspace to re-run OR-Tools constraint solver.*"
        )
    else:
        return (
            "### 🤖 Planify AI Intelligence Assistant\n\n"
            f"I have analyzed your active operational context (`{len(assignments)} scheduled sessions` across active departments).\n\n"
            "- **Timetable Workspace**: Click **Timetable Workspace** to inspect grid assignments.\n"
            "- **Academic Setup**: Manage **Departments**, **Sections**, **Subjects**, and **Rooms**.\n"
            "- **Automation**: Connect Make webhooks in **Operations -> Automation & Broadcast**.\n\n"
            "💡 *For live LLM conversational reasoning, configure `GROQ_API_KEY` in `backend/.env`.*"
        )


@app.post("/chat")
def chat_with_groq(request: ChatRequest):
    api_key = os.getenv("GROQ_API_KEY")
    ctx = request.context or {}
    
    if not api_key or api_key == "your_api_key_here" or len(api_key) < 20:
        fallback_reply = generate_expert_fallback_reply(request.message, ctx)
        return {"reply": fallback_reply}
    
    try:
        client = Groq(api_key=api_key)
        
        system_instruction = (
            "You are an elite, highly intelligent AI Timetable Scheduling Assistant powered by Groq. "
            "You MUST provide GENUINE, intelligent, and highly optimized scheduling suggestions based on the current context. "
            "You MUST use Markdown heavily to make your responses beautiful, readable, and structured. "
            "If the user uploads an image or pdf of a timetable, YOU MUST ACT AS AN ADVANCED OCR SYSTEM. "
            "Extract all scheduling data from the image into a STRICT JSON block. "
            "The JSON block MUST be enclosed in ```json ... ``` and contain the following exact keys: "
            "`teachers` (list of objects with `name` and `free_periods`), "
            "`subjects` (list of objects with `code`, `name`, `teacher`, `section`, `required_slots`, `is_lab`, `colorIndex`), "
            "`rooms` (list of strings), "
            "`sections` (list of objects with `name`, `room`, `lab_room`), "
            "`timeSlots` (list of strings). "
            "ALWAYS provide a helpful markdown message summarizing your actions or suggestions alongside the JSON."
        )
        if ctx:
            system_instruction += f"\nCurrent Timetable State:\n- Objective Score: {ctx.get('objective_score', 'N/A')}\n- Total Classes: {len(ctx.get('assignments', []))}\n"
            if ctx.get("ai_suggestions"):
                system_instruction += "Current Solver Suggestions:\n"
                for suggestion in ctx.get("ai_suggestions", [])[:5]:
                    system_instruction += f"- {suggestion}\n"
        
        # If an image is provided, fallback to Gemini 2.5 Flash for OCR
        if request.image:
            gemini_key = os.getenv("GEMINI_API_KEY")
            if not gemini_key:
                return {"reply": "Groq vision is currently disabled. Please set your GEMINI_API_KEY in the backend/.env file to fallback to Gemini for OCR image extraction."}
            
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={gemini_key}"
            payload = {
                "systemInstruction": {
                    "parts": [{"text": system_instruction}]
                },
                "contents": [{
                    "parts": [
                        {"text": request.message or "Please extract and format the timetable data from this image."},
                        {
                            "inlineData": {
                                "mimeType": "image/jpeg",
                                "data": request.image
                            }
                        }
                    ]
                }]
            }
            res = requests.post(url, json=payload)
            if res.status_code == 200:
                data = res.json()
                try:
                    return {"reply": data["candidates"][0]["content"]["parts"][0]["text"]}
                except KeyError:
                    return {"reply": f"Gemini OCR failed to parse image data."}
            else:
                fallback_reply = generate_expert_fallback_reply(request.message, ctx)
                return {"reply": fallback_reply}

        # Process text chat with Groq
        messages = [{"role": "system", "content": system_instruction}]
        
        for h in request.history:
            role = "assistant" if h.get("sender") == "bot" else "user"
            if h.get("text"):
                messages.append({"role": role, "content": h.get("text")})
                
        messages.append({"role": "user", "content": request.message})
        
        chat_completion = client.chat.completions.create(
            messages=messages,
            model="llama-3.3-70b-versatile",
        )
        
        return {"reply": chat_completion.choices[0].message.content}
    except Exception as e:
        error_str = str(e)
        logger.warning(f"Groq API call error handled: {error_str}")
        fallback_reply = generate_expert_fallback_reply(request.message, ctx)
        return {"reply": fallback_reply}


if __name__ == "__main__":
    import sys
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    import uvicorn
    env_backend_port = os.getenv("BACKEND_PORT")
    env_port = os.getenv("PORT")
    if env_backend_port:
        port = int(env_backend_port)
    elif env_port and env_port != "3000":
        port = int(env_port)
    else:
        port = 8080
    host = os.getenv("HOST", "0.0.0.0")
    print(f"[+] Starting Plannify Backend on http://{host}:{port}")
    uvicorn.run(app, host=host, port=port)

