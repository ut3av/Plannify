"""
Stylized OpenpyXL Excel Timetable Exporter.
Generates publication-ready institutional schedules for individual teachers and full institutional master rosters.
"""
import io
import base64
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri"]


def create_teacher_excel(teacher_name: str, slots: List[str], assignments: List[dict]) -> str:
    """
    Generates a base64-encoded Excel file for a specific teacher's timetable.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = (teacher_name[:31]) if teacher_name else "Teacher"

    # Define Styles
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    sub_header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 1. Header Info Row
    ws.append([f"Faculty Schedule: {teacher_name}", "", "", "", "", "", f"Generated: {datetime.now().strftime('%d/%m/%Y')}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    # 2. Period & Time Rows
    period_nums = ["Day / Period"]
    time_slots_row = [""]
    lunch_cols = []

    period_counter = 1
    roman_numerals = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"]

    for i, slot in enumerate(slots):
        p_num = roman_numerals[period_counter - 1] if period_counter <= len(roman_numerals) else str(period_counter)
        period_nums.append(p_num)
        time_slots_row.append(slot)

        if i < len(slots) - 1:
            try:
                end_match = slot.split("-")[1].strip()
                next_start_match = slots[i + 1].split("-")[0].strip()
                if end_match != next_start_match:
                    period_nums.append("")
                    time_slots_row.append("LUNCH")
                    lunch_cols.append(len(time_slots_row))
            except (IndexError, AttributeError):
                pass
        period_counter += 1

    ws.append(period_nums)
    ws.append(time_slots_row)

    # Apply styling to headers
    for r in [2, 3]:
        for c in range(1, len(time_slots_row) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(bold=True)
            cell.fill = sub_header_fill
            cell.border = border
            cell.alignment = alignment

    # 3. Data Rows
    for day in DAYS:
        row_data = [day]
        slot_idx = 0
        for i in range(1, len(time_slots_row)):
            col_idx = i + 1
            if col_idx in lunch_cols:
                row_data.append("")
                continue

            slot_name = slots[slot_idx] if slot_idx < len(slots) else ""
            match = next((a for a in assignments if a.get("teacher") == teacher_name and a.get("day") == day and a.get("slot") == slot_name), None)

            if match:
                code_display = match.get("code") or match.get("subject")
                val = f"{code_display}\n({match.get('room')})\n[{match.get('section') or 'Auto'}]"
                row_data.append(val)
            else:
                row_data.append("")
            slot_idx += 1
        ws.append(row_data)

    # Style grid cells
    for r in range(4, 4 + len(DAYS)):
        for c in range(1, len(time_slots_row) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = alignment
            if c == 1:
                cell.font = Font(bold=True)
                cell.fill = sub_header_fill

    # Handle vertical Lunch merges
    for col_idx in lunch_cols:
        ws.merge_cells(start_row=3, start_column=col_idx, end_row=3 + len(DAYS), end_column=col_idx)
        cell = ws.cell(row=3, column=col_idx)
        cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)

    # Column widths
    ws.column_dimensions["A"].width = 15
    for c in range(2, len(time_slots_row) + 1):
        char = chr(64 + c) if c <= 26 else f"{chr(64 + (c - 1) // 26)}{chr(64 + (c - 1) % 26 + 1)}"
        ws.column_dimensions[char].width = 18

    output = io.BytesIO()
    wb.save(output)
    return base64.b64encode(output.getvalue()).decode("utf-8")


def create_master_timetable_excel(slots: List[str], assignments: List[dict], sections: List[str]) -> str:
    """
    Generates a master workbook containing a tab for each academic section and a Master Overview.
    """
    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    for section in (sections or ["General"]):
        ws = wb.create_sheet(title=f"Section {section}"[:31])
        ws.append([f"Section Schedule: {section}", "", "", "", "", "", f"Exported: {datetime.now().strftime('%d/%m/%Y')}"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(slots) + 1)

        ws.append(["Day / Slot", *slots])
        for day in DAYS:
            row = [day]
            for slot in slots:
                matches = [a for a in assignments if a.get("section") == section and a.get("day") == day and a.get("slot") == slot]
                if matches:
                    cell_text = "\n".join(f"{m.get('subject')} ({m.get('teacher')}) [{m.get('room')}]" for m in matches)
                    row.append(cell_text)
                else:
                    row.append("—")
            ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    return base64.b64encode(output.getvalue()).decode("utf-8")
