"""
FastAPI Analytics API Routes for Planify.exe.
Prefix: /analytics
"""

import io
import os
from typing import Optional, List
from fastapi import APIRouter, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from groq import Groq
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from .analytics_db import (
        get_dashboard_kpis,
        get_faculty_directory_analytics,
        get_individual_faculty_analytics,
        get_department_analytics,
        get_operational_insights,
        get_workload_config,
        update_workload_config,
        seed_30day_demo_history,
        clear_all_demo_data,
    )
except ImportError:
    from analytics_db import (
        get_dashboard_kpis,
        get_faculty_directory_analytics,
        get_individual_faculty_analytics,
        get_department_analytics,
        get_operational_insights,
        get_workload_config,
        update_workload_config,
        seed_30day_demo_history,
        clear_all_demo_data,
    )

router = APIRouter(prefix="/analytics", tags=["Faculty Analytics"])


@router.post("/seed-demo-history")
def seed_analytics_demo_history():
    """Seeds 30 days of rich LNCT operational history (attendance, half days, substitutions, leaves)."""
    try:
        return seed_30day_demo_history()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/clear-demo")
@router.delete("/clear-demo")
def clear_analytics_demo_history():
    """Purges demo attendance punches, substitution logs, and leave records."""
    try:
        return clear_all_demo_data()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



class WorkloadConfigInput(BaseModel):
    low_threshold: Optional[int] = Field(default=12, ge=1, le=40)
    high_threshold: Optional[int] = Field(default=18, ge=1, le=60)
    academic_year: Optional[str] = "2026-27"


class AnalyticsChatRequest(BaseModel):
    message: str
    faculty_id: Optional[str] = None
    department_id: Optional[str] = None
    range_key: Optional[str] = "30d"
    history: List[dict] = []


# ── Dashboard & KPIs ──────────────────────────────────────────────────────────

@router.get("/dashboard")
def get_analytics_dashboard(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    range_key: Optional[str] = Query("30d"),
    department_id: Optional[str] = Query(None),
):
    try:
        return get_dashboard_kpis(start_date=start_date, end_date=end_date, range_key=range_key, department_id=department_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Faculty Directory Analytics ───────────────────────────────────────────────

@router.get("/faculty")
def get_faculty_directory_table(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    range_key: Optional[str] = Query("30d"),
    department_id: Optional[str] = Query(None),
    designation: Optional[str] = Query(None),
    employment_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    att_min: Optional[float] = Query(None),
    att_max: Optional[float] = Query(None),
    sort_by: Optional[str] = Query("teacher_name"),
    sort_order: Optional[str] = Query("asc"),
):
    try:
        return get_faculty_directory_analytics(
            start_date=start_date,
            end_date=end_date,
            range_key=range_key,
            department_id=department_id,
            designation=designation,
            employment_type=employment_type,
            status=status,
            search=search,
            att_min=att_min,
            att_max=att_max,
            sort_by=sort_by,
            sort_order=sort_order,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Individual Faculty Profile Analytics ──────────────────────────────────────

@router.get("/faculty/{faculty_id}")
def get_faculty_profile_analytics(
    faculty_id: str,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    range_key: Optional[str] = Query("30d"),
):
    try:
        res = get_individual_faculty_analytics(faculty_id, start_date=start_date, end_date=end_date, range_key=range_key)
        if not res:
            raise HTTPException(status_code=404, detail="Faculty profile not found")
        return res
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/faculty/{faculty_id}/attendance")
def get_faculty_attendance_analytics(
    faculty_id: str,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    range_key: Optional[str] = Query("30d"),
):
    profile = get_individual_faculty_analytics(faculty_id, start_date=start_date, end_date=end_date, range_key=range_key)
    if not profile:
        raise HTTPException(status_code=404, detail="Faculty profile not found")
    return profile.get("attendance", {})


@router.get("/faculty/{faculty_id}/leave")
def get_faculty_leave_analytics(
    faculty_id: str,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    range_key: Optional[str] = Query("30d"),
):
    profile = get_individual_faculty_analytics(faculty_id, start_date=start_date, end_date=end_date, range_key=range_key)
    if not profile:
        raise HTTPException(status_code=404, detail="Faculty profile not found")
    return profile.get("leave", {})


@router.get("/faculty/{faculty_id}/substitutions")
def get_faculty_substitution_analytics(
    faculty_id: str,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    range_key: Optional[str] = Query("30d"),
):
    profile = get_individual_faculty_analytics(faculty_id, start_date=start_date, end_date=end_date, range_key=range_key)
    if not profile:
        raise HTTPException(status_code=404, detail="Faculty profile not found")
    return profile.get("substitution", {})


@router.get("/faculty/{faculty_id}/workload")
def get_faculty_workload_analytics(
    faculty_id: str,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    range_key: Optional[str] = Query("30d"),
):
    profile = get_individual_faculty_analytics(faculty_id, start_date=start_date, end_date=end_date, range_key=range_key)
    if not profile:
        raise HTTPException(status_code=404, detail="Faculty profile not found")
    return profile.get("workload", {})


# ── Department Analytics ──────────────────────────────────────────────────────

@router.get("/departments")
def get_departments_analytics(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    range_key: Optional[str] = Query("30d"),
):
    try:
        return get_department_analytics(start_date=start_date, end_date=end_date, range_key=range_key)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Operational Insights & Early Warnings ─────────────────────────────────────

@router.get("/insights")
def get_insights(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    range_key: Optional[str] = Query("30d"),
):
    try:
        return get_operational_insights(start_date=start_date, end_date=end_date, range_key=range_key)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Institutional Workload Configuration ──────────────────────────────────────

@router.get("/config")
def get_config():
    return get_workload_config()


@router.put("/config")
def set_config(input_data: WorkloadConfigInput):
    return update_workload_config(input_data.model_dump(exclude_none=True))


# ── AI Analytics Assistant ────────────────────────────────────────────────────

@router.post("/ai-chat")
def analytics_ai_chat(request: AnalyticsChatRequest):
    api_key = os.getenv("GROQ_API_KEY")
    if not api_key or api_key == "your_api_key_here":
        return {"reply": "Please set your GROQ_API_KEY in the backend/.env file to use the AI Analytics Assistant."}

    try:
        # Fetch current verified DB analytics context
        kpis = get_dashboard_kpis(range_key=request.range_key, department_id=request.department_id)
        insights = get_operational_insights(range_key=request.range_key)
        fac_list = get_faculty_directory_analytics(range_key=request.range_key, department_id=request.department_id)

        context_summary = f"""
VERIFIED INSTITUTIONAL ANALYTICS CONTEXT ({kpis['time_period']['formatted']}):
- Total Faculty: {kpis['kpis']['total_faculty']['value']}
- Attendance Rate: {kpis['kpis']['attendance_rate']['value']}%
- Punctuality Rate: {kpis['kpis']['average_punctuality']['value']}%
- Total Leave Days: {kpis['kpis']['leave_days']['value']}
- Conducted Classes: {kpis['kpis']['classes_conducted']['value']}
- Substitutions Provided: {kpis['kpis']['substitutions_provided']['value']}
- Substitutions Received: {kpis['kpis']['substitutions_received']['value']}
- Schedule Changes: {kpis['kpis']['schedule_changes']['value']}

OPERATIONAL INSIGHTS SUMMARY:
"""
        for ins in insights:
            context_summary += f"- [{ins['severity'].upper()}] {ins['title']}: {ins['message']}\n"

        context_summary += "\nFACULTY HIGHLIGHTS:\n"
        for f in fac_list[:8]:
            context_summary += f"- {f['teacher_name']} ({f['department_name']}, {f['designation']}): Att {f['attendance_percentage']}%, Workload {f['weekly_workload']} periods/wk ({f['workload_status']}), Leave {f['leave_days']} days, Subs {f['substitutions_provided']} prov.\n"

        system_prompt = (
            "You are an expert, professional Academic Operations & Faculty Analytics Assistant for Planify.exe. "
            "IMPORTANT RULES:\n"
            "1. You MUST NEVER invent or fabricate statistics, numbers, or performance grades.\n"
            "2. Base all explanations strictly on the provided verified institutional database context above.\n"
            "3. Explicitly separate operational scheduling facts from academic teaching quality.\n"
            "4. NEVER penalize or judge a faculty member for legitimate medical leave or approved leaves.\n"
            "5. Format your answers clearly using Markdown headers, bullet points, and key callouts."
        )

        messages = [
            {"role": "system", "content": system_prompt + "\n\n" + context_summary},
        ]
        for h in request.history:
            role = "assistant" if h.get("sender") == "bot" else "user"
            if h.get("text"):
                messages.append({"role": role, "content": h.get("text")})

        messages.append({"role": "user", "content": request.message})

        client = Groq(api_key=api_key)
        chat_completion = client.chat.completions.create(
            messages=messages,
            model="llama-3.3-70b-versatile",
        )
        return {"reply": chat_completion.choices[0].message.content}
    except Exception as e:
        return {"reply": f"Error communicating with AI Analytics Assistant: {str(e)}"}


# ── Reports Export Endpoint ───────────────────────────────────────────────────

@router.get("/export")
def export_analytics_report(
    report_type: str = Query("faculty_summary"),
    range_key: str = Query("30d"),
    format_type: str = Query("excel"),
    department_id: Optional[str] = Query(None),
):
    try:
        fac_analytics = get_faculty_directory_analytics(range_key=range_key, department_id=department_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Faculty Operations"

        # Headers
        headers = [
            "Employee ID", "Faculty Name", "Department", "Designation",
            "Attendance %", "Present Days", "Late Days", "Leave Days",
            "Classes Conducted", "Subs Provided", "Subs Received", "Weekly Load", "Workload Status"
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, f in enumerate(fac_analytics, 2):
            row = [
                f["employee_id"], f["teacher_name"], f["department_name"], f["designation"],
                f["attendance_percentage"], f["present_days"], f["late_days"], f["leave_days"],
                f["classes_conducted"], f["substitutions_provided"], f["substitutions_received"],
                f["weekly_workload"], f["workload_status"]
            ]
            ws.append(row)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Planify_Faculty_Analytics_{range_key}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
