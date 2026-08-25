"""
Document Ingestion & Parser Service for Planify.exe.
Parses multi-page institutional PDFs, multi-sheet Excel workbooks, and document scans.
"""
import io
import os
import re
import logging
from typing import Dict, List, Any, Optional, Tuple

import openpyxl
from pypdf import PdfReader
import pdfplumber

from .data_normalizer import (
    normalize_whitespace,
    normalize_faculty_name,
    normalize_subject_code,
    normalize_subject_name,
    parse_program_and_section,
    normalize_room_number,
    normalize_phone_number
)

logger = logging.getLogger("plannify.parser")


class DocumentParser:
    def __init__(self):
        pass

    # ─────────────────────────────────────────────────────────────
    # 1. PDF Parsing Engine
    # ─────────────────────────────────────────────────────────────

    def parse_pdf(self, file_bytes: bytes, filename: str = "document.pdf") -> Dict[str, Any]:
        """
        Parses a PDF file containing institutional timetable and allocation records.
        Deterministically extracts section blocks, subjects, faculty names, mentors, and contact details.
        """
        result = {
            "source_file": filename,
            "file_type": "pdf",
            "sections": [],
            "faculty": [],
            "subjects": [],
            "allocations": [],
            "rooms": [],
            "mentors": [],
            "raw_extractions": []
        }

        try:
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                full_text = ""
                all_tables = []
                for page_idx, page in enumerate(pdf.pages):
                    page_text = page.extract_text() or ""
                    full_text += f"\n--- Page {page_idx+1} ---\n" + page_text
                    tables = page.extract_tables()
                    for t in tables:
                        all_tables.append({"page": page_idx + 1, "table": t})

            # Check if PDF contains Section headers like "BCA-I Sem. Section A (AI/DA)"
            section_matches = list(re.finditer(r'([A-Za-z0-9\-_]+-I{1,3}\s+Sem\.\s+Section\s+[A-G](?:\s*\([A-Za-z0-9\/\s\-]+\))?)', full_text))
            
            # Extract mentors from tables and text patterns on pages 3 & 4
            mentor_list = []
            seen_phones = set()

            # Method A: From tables on pages 3 & 4
            for item in all_tables:
                table = item["table"]
                if not table or len(table) < 2:
                    continue
                # Check for phone numbers in table cells
                for r_idx in range(len(table)):
                    row = table[r_idx]
                    for c_idx, cell in enumerate(row):
                        cell_str = normalize_whitespace(str(cell or ""))
                        digits = re.sub(r'\D', '', cell_str)
                        if len(digits) == 10 and digits not in seen_phones:
                            # Name is in preceding row or column
                            m_name = None
                            if r_idx > 0 and table[r_idx - 1][c_idx]:
                                m_name = normalize_whitespace(str(table[r_idx - 1][c_idx]))
                            elif r_idx > 1 and table[r_idx - 2][c_idx]:
                                m_name = normalize_whitespace(str(table[r_idx - 2][c_idx]))

                            if m_name and "library" not in m_name.lower() and "mentor" not in m_name.lower():
                                m_name_norm = normalize_faculty_name(m_name)["normalized_name"]
                                mentor_list.append({"name": m_name_norm, "phone": normalize_phone_number(digits)})
                                seen_phones.add(digits)

            # Method B: Regex fallback for any remaining phone numbers
            for match in re.finditer(r'([A-Za-z\.\s]{3,30})\n\s*(\d{10})\b', full_text):
                raw_name = normalize_whitespace(match.group(1))
                phone = match.group(2)
                if phone not in seen_phones and "library" not in raw_name.lower() and "mentor" not in raw_name.lower():
                    m_norm = normalize_faculty_name(raw_name)["normalized_name"]
                    mentor_list.append({"name": m_norm, "phone": normalize_phone_number(phone)})
                    seen_phones.add(phone)

            # Parse Table Grids
            parsed_sections = []
            parsed_allocations = []
            parsed_faculty_set = {}
            parsed_subjects_set = {}

            # Stitch tables that split across page boundaries
            merged_tables = []
            pending_header = None

            for item in all_tables:
                table = item["table"]
                if not table:
                    continue
                first_cell = normalize_whitespace(str(table[0][0] or "")).lower()

                if "subject code" in first_cell or "sub code" in first_cell:
                    if len(table) == 1:
                        # Split table header at bottom of page
                        pending_header = item
                    else:
                        merged_tables.append(item)
                        pending_header = None
                elif "subject name" in first_cell or "sub name" in first_cell:
                    if pending_header:
                        # Merge with pending header from previous page
                        stitched = {
                            "page": pending_header["page"],
                            "table": pending_header["table"] + table
                        }
                        merged_tables.append(stitched)
                        pending_header = None
                    else:
                        merged_tables.append(item)
                else:
                    merged_tables.append(item)

            # Iterate through merged tables
            current_sec_idx = 0
            for item in merged_tables:
                table = item["table"]
                if not table or len(table) < 2:
                    continue

                # Check if this table has Subject Code row
                header_row = table[0]
                first_cell = normalize_whitespace(str(header_row[0] or "")).lower()

                if "subject code" in first_cell or "sub code" in first_cell:
                    # Find matching section name from text if available
                    sec_info = None
                    if current_sec_idx < len(section_matches):
                        sec_raw = section_matches[current_sec_idx].group(1)
                        sec_info = parse_program_and_section(sec_raw)
                        current_sec_idx += 1
                    else:
                        sec_info = parse_program_and_section(f"BCA-I Sem. Section {chr(65 + len(parsed_sections))}")

                    # Attach mentor if available
                    if len(parsed_sections) < len(mentor_list):
                        m_obj = mentor_list[len(parsed_sections)]
                        sec_info["mentor_name"] = m_obj["name"]
                        sec_info["mentor_phone"] = m_obj["phone"]
                    else:
                        sec_info["mentor_name"] = None
                        sec_info["mentor_phone"] = None

                    parsed_sections.append(sec_info)

                    # Extract Subject Codes from Row 0
                    codes = [normalize_subject_code(str(c or "")) for c in header_row[1:] if c and "library" not in str(c).lower() and "mentor" not in str(c).lower()]
                    
                    # Extract Subject Names from Row 1
                    names_row = table[1] if len(table) > 1 else []
                    names = [normalize_subject_name(str(c or "")) for c in names_row[1:len(codes)+1]] if len(names_row) > 1 else []

                    # Extract Faculty Names from Row 2
                    fac_row = table[2] if len(table) > 2 else []
                    facs = [normalize_whitespace(str(c or "")) for c in fac_row[1:len(codes)+1]] if len(fac_row) > 2 else []

                    for idx, code in enumerate(codes):
                        if not code or code in ["LIBRARY", "MENTOR"]:
                            continue
                        sub_name = names[idx] if idx < len(names) and names[idx] else f"Course {code}"
                        raw_fac = facs[idx] if idx < len(facs) and facs[idx] else "TBA Faculty"
                        fac_norm = normalize_faculty_name(raw_fac)

                        is_lab = "lab" in sub_name.lower() or "lab" in code.lower()

                        # Collect subject
                        sub_key = f"{sec_info['program_code']}_{code}"
                        if sub_key not in parsed_subjects_set:
                            parsed_subjects_set[sub_key] = {
                                "code": code,
                                "name": sub_name,
                                "program_code": sec_info["program_code"],
                                "semester_number": sec_info["semester_number"],
                                "is_lab": is_lab,
                                "credit_hours": 4.0 if not is_lab else 2.0,
                                "lecture_hours": 4 if not is_lab else 0,
                                "lab_hours": 4 if is_lab else 0,
                            }

                        # Collect faculty
                        if fac_norm["normalized_name"] and "new faculty" not in fac_norm["normalized_name"].lower() and "tba" not in fac_norm["normalized_name"].lower():
                            f_key = fac_norm["canonical_name"].lower()
                            if f_key not in parsed_faculty_set:
                                parsed_faculty_set[f_key] = {
                                    "teacher_name": fac_norm["normalized_name"],
                                    "canonical_name": fac_norm["canonical_name"],
                                    "designation": "Assistant Professor",
                                    "department": "Computer Applications",
                                    "phone": sec_info.get("mentor_phone") if sec_info.get("mentor_name") == fac_norm["normalized_name"] else None
                                }

                        # Collect allocation
                        parsed_allocations.append({
                            "program_code": sec_info["program_code"],
                            "semester_number": sec_info["semester_number"],
                            "section_name": sec_info["section_name"],
                            "section_full_name": sec_info["full_name"],
                            "subject_code": code,
                            "subject_name": sub_name,
                            "faculty_name": fac_norm["normalized_name"] if fac_norm["normalized_name"] else "TBA Faculty",
                            "canonical_faculty_name": fac_norm["canonical_name"],
                            "is_lab": is_lab,
                            "weekly_load": 4 if not is_lab else 2,
                            "source_location": f"Page {item['page']}, Section {sec_info['section_letter']}"
                        })

            result["sections"] = parsed_sections
            result["subjects"] = list(parsed_subjects_set.values())
            result["faculty"] = list(parsed_faculty_set.values())
            result["allocations"] = parsed_allocations
            result["mentors"] = mentor_list

            logger.info(f"[PDF Parser] Extracted {len(parsed_sections)} sections, {len(parsed_subjects_set)} subjects, {len(parsed_faculty_set)} faculty, {len(parsed_allocations)} allocations from {filename}.")
        except Exception as e:
            logger.error(f"[PDF Parser Error] {e}", exc_info=True)
            result["error"] = str(e)

        return result

    # ─────────────────────────────────────────────────────────────
    # 2. Excel Parsing Engine
    # ─────────────────────────────────────────────────────────────

    def parse_excel(self, file_bytes: bytes, filename: str = "workbook.xlsx") -> Dict[str, Any]:
        """
        Parses multi-sheet Excel workbooks (`Sub-AllOC`, `Teaching Load`, `Projector Room`, `Lab Time Table`, `Room Shifting`, etc.).
        Extracts sections, faculty allocations, workloads, room capabilities, and facility shifts.
        """
        result = {
            "source_file": filename,
            "file_type": "xlsx",
            "sheets_detected": [],
            "sections": [],
            "faculty": [],
            "subjects": [],
            "allocations": [],
            "rooms": [],
            "teaching_loads": [],
            "room_shifts": [],
            "mentors": [],
        }

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            result["sheets_detected"] = wb.sheetnames

            # 1. Parse Sub-AllOC / Sheet1 (Sections & Allocations)
            target_alloc_sheet = None
            for sname in ["Sub-AllOC", "Sheet1", "Subject Allocation", "Allocation"]:
                if sname in wb.sheetnames:
                    target_alloc_sheet = wb[sname]
                    break

            if target_alloc_sheet:
                self._parse_sub_alloc_sheet(target_alloc_sheet, result)

            # 2. Parse Teaching Load sheet if present
            if "Teaching Load" in wb.sheetnames:
                self._parse_teaching_load_sheet(wb["Teaching Load"], result)

            # 3. Parse Projector Room sheet if present
            if "Projector Room" in wb.sheetnames:
                self._parse_projector_room_sheet(wb["Projector Room"], result)

            # 4. Parse Room Shifting sheet if present
            if "Room Shifting" in wb.sheetnames:
                self._parse_room_shifting_sheet(wb["Room Shifting"], result)

            # 5. Parse Lab Time Table if present
            for lsheet in ["Lab Time Table", "BTech Lab"]:
                if lsheet in wb.sheetnames:
                    self._parse_lab_timetable_sheet(wb[lsheet], result)

            logger.info(f"[Excel Parser] Processed {filename} across {len(wb.sheetnames)} sheets. Found {len(result['sections'])} sections, {len(result['faculty'])} faculty, {len(result['subjects'])} subjects, {len(result['rooms'])} rooms.")
        except Exception as e:
            logger.error(f"[Excel Parser Error] {e}", exc_info=True)
            result["error"] = str(e)

        return result

    def _parse_sub_alloc_sheet(self, sheet, result: Dict[str, Any]):
        """Extracts section blocks and faculty allocations from Sub-AllOC or Sheet1."""
        rows = list(sheet.iter_rows(values_only=True))
        r_idx = 0
        total_rows = len(rows)

        parsed_sections = []
        parsed_allocations = []
        parsed_faculty_dict = {}
        parsed_subjects_dict = {}

        while r_idx < total_rows:
            row = rows[r_idx]
            if not row or not any(row):
                r_idx += 1
                continue

            first_cell = normalize_whitespace(str(row[0] or ""))

            # Check for Section Header: e.g. "BCA-I Sem. Section A (AI/DA)" or "MCA-I Sem. Section D"
            is_section_header = (
                ("sem." in first_cell.lower() or "semester" in first_cell.lower()) and
                ("section" in first_cell.lower() or "sec" in first_cell.lower())
            )

            if is_section_header:
                sec_info = parse_program_and_section(first_cell)
                sec_row_idx = r_idx

                # Next row is usually Subject Code
                codes_row = rows[r_idx + 1] if r_idx + 1 < total_rows else None
                names_row = rows[r_idx + 2] if r_idx + 2 < total_rows else None
                fac_row = rows[r_idx + 3] if r_idx + 3 < total_rows else None

                if codes_row and len(codes_row) > 1 and "subject code" in normalize_whitespace(str(codes_row[0] or "")).lower():
                    # Extract Mentor if present in codes_row or names_row or fac_row
                    mentor_name = None
                    mentor_phone = None

                    # Mentor col index is usually last column
                    for col_idx in range(len(codes_row)):
                        cell_val = normalize_whitespace(str(codes_row[col_idx] or "")).lower()
                        if "mentor" in cell_val:
                            if names_row and col_idx < len(names_row) and names_row[col_idx]:
                                mentor_name = normalize_whitespace(str(names_row[col_idx]))
                            if fac_row and col_idx < len(fac_row) and fac_row[col_idx]:
                                mentor_phone = normalize_phone_number(str(fac_row[col_idx]))

                    sec_info["mentor_name"] = mentor_name
                    sec_info["mentor_phone"] = mentor_phone
                    parsed_sections.append(sec_info)

                    # Extract codes, names, faculty
                    for c_idx in range(1, len(codes_row)):
                        code_val = normalize_subject_code(str(codes_row[c_idx] or ""))
                        if not code_val or code_val in ["LIBRARY", "MENTOR", "TOTAL", "NONE"]:
                            continue

                        name_val = normalize_subject_name(str(names_row[c_idx] or "")) if names_row and c_idx < len(names_row) else f"Course {code_val}"
                        raw_fac = normalize_whitespace(str(fac_row[c_idx] or "")) if fac_row and c_idx < len(fac_row) else "TBA Faculty"
                        fac_norm = normalize_faculty_name(raw_fac)

                        is_lab = "lab" in name_val.lower() or "lab" in code_val.lower()

                        # Collect subject
                        sub_key = f"{sec_info['program_code']}_{code_val}"
                        if sub_key not in parsed_subjects_dict:
                            parsed_subjects_dict[sub_key] = {
                                "code": code_val,
                                "name": name_val,
                                "program_code": sec_info["program_code"],
                                "semester_number": sec_info["semester_number"],
                                "is_lab": is_lab,
                                "credit_hours": 4.0 if not is_lab else 2.0,
                                "lecture_hours": 4 if not is_lab else 0,
                                "lab_hours": 4 if is_lab else 0,
                            }

                        # Collect faculty
                        if fac_norm["normalized_name"] and "new faculty" not in fac_norm["normalized_name"].lower() and "tba" not in fac_norm["normalized_name"].lower():
                            f_key = fac_norm["canonical_name"].lower()
                            if f_key not in parsed_faculty_dict:
                                parsed_faculty_dict[f_key] = {
                                    "teacher_name": fac_norm["normalized_name"],
                                    "canonical_name": fac_norm["canonical_name"],
                                    "designation": "Assistant Professor",
                                    "department": "Computer Applications",
                                    "phone": mentor_phone if mentor_name and fac_norm["canonical_name"].lower() in mentor_name.lower() else None
                                }

                        # Collect allocation
                        parsed_allocations.append({
                            "program_code": sec_info["program_code"],
                            "semester_number": sec_info["semester_number"],
                            "section_name": sec_info["section_name"],
                            "section_full_name": sec_info["full_name"],
                            "subject_code": code_val,
                            "subject_name": name_val,
                            "faculty_name": fac_norm["normalized_name"] if fac_norm["normalized_name"] else "TBA Faculty",
                            "canonical_faculty_name": fac_norm["canonical_name"],
                            "is_lab": is_lab,
                            "weekly_load": 4 if not is_lab else 2,
                            "source_location": f"Sheet: {sheet.title}, Row: {sec_row_idx + 1}"
                        })

                    r_idx += 3
            r_idx += 1

        result["sections"].extend(parsed_sections)
        result["subjects"].extend(list(parsed_subjects_dict.values()))
        result["faculty"].extend(list(parsed_faculty_dict.values()))
        result["allocations"].extend(parsed_allocations)

    def _parse_teaching_load_sheet(self, sheet, result: Dict[str, Any]):
        """Parses faculty weekly teaching load and splits."""
        rows = list(sheet.iter_rows(values_only=True))
        loads = []
        for r_idx, row in enumerate(rows):
            if not row or len(row) < 3:
                continue
            first = normalize_whitespace(str(row[0] or ""))
            second = normalize_whitespace(str(row[1] or ""))

            # Skip header
            if "faculty name" in second.lower() or "session" in first.lower() or not second:
                continue

            fac_norm = normalize_faculty_name(second)
            if not fac_norm["normalized_name"]:
                continue

            # Check total load column (often last column or column 8)
            total_load = 18
            for c in reversed(row):
                if c is not None:
                    try:
                        val = float(c)
                        if 1.0 <= val <= 40.0:
                            total_load = int(val)
                            break
                    except Exception:
                        pass

            loads.append({
                "faculty_name": fac_norm["normalized_name"],
                "canonical_name": fac_norm["canonical_name"],
                "total_load": total_load,
                "row": r_idx + 1
            })

            # Update existing faculty profile load if present
            for f in result["faculty"]:
                if f["canonical_name"].lower() == fac_norm["canonical_name"].lower():
                    f["max_weekly_hours"] = total_load

        result["teaching_loads"] = loads

    def _parse_projector_room_sheet(self, sheet, result: Dict[str, Any]):
        """Parses classrooms with Projector & Smart LED panel capabilities."""
        rows = list(sheet.iter_rows(values_only=True))
        rooms_dict = {}
        for row in rows:
            if not row or len(row) < 3:
                continue
            r_cell = normalize_whitespace(str(row[1] or ""))
            class_cell = normalize_whitespace(str(row[2] or ""))

            if "class room" in r_cell.lower() or not r_cell:
                continue

            room_info = normalize_room_number(r_cell)
            if not room_info["room_number"]:
                continue

            has_proj = any("yes" in normalize_whitespace(str(c or "")).lower() for c in row[3:5])
            has_smart = any("smart" in normalize_whitespace(str(c or "")).lower() or "led" in normalize_whitespace(str(c or "")).lower() for c in row[3:5])

            room_key = room_info["room_number"]
            if room_key not in rooms_dict:
                rooms_dict[room_key] = {
                    "room_number": room_info["room_number"],
                    "room_type": room_info["room_type"],
                    "capacity": 60,
                    "has_projector": has_proj,
                    "has_smart_board": has_smart,
                    "building_name": "Ramnath Guha Block",
                    "associated_class": class_cell if class_cell else None
                }

        result["rooms"].extend(list(rooms_dict.values()))

    def _parse_room_shifting_sheet(self, sheet, result: Dict[str, Any]):
        """Parses classroom shifts from MCA Building to Agriculture Building."""
        rows = list(sheet.iter_rows(values_only=True))
        shifts = []
        for row in rows:
            if not row or len(row) < 4:
                continue
            first = normalize_whitespace(str(row[0] or ""))
            class_name = normalize_whitespace(str(row[1] or ""))
            from_room = normalize_whitespace(str(row[2] or ""))
            to_room = normalize_whitespace(str(row[3] or ""))

            if "class room from" in from_room.lower() or not class_name or not from_room:
                continue

            shifts.append({
                "class": class_name,
                "from_room": normalize_room_number(from_room)["room_number"],
                "to_room": normalize_room_number(to_room)["room_number"],
                "from_building": "LNCT (MCA) Building",
                "to_building": "Agriculture Building"
            })

            # Add to_room to rooms list
            to_r = normalize_room_number(to_room)["room_number"]
            if to_r and not any(r["room_number"] == to_r for r in result["rooms"]):
                result["rooms"].append({
                    "room_number": to_r,
                    "room_type": "CLASSROOM",
                    "capacity": 60,
                    "has_projector": False,
                    "has_smart_board": False,
                    "building_name": "Agriculture Building"
                })

        result["room_shifts"] = shifts

    def _parse_lab_timetable_sheet(self, sheet, result: Dict[str, Any]):
        """Parses dedicated computer labs (LAB-002, LAB-003)."""
        rows = list(sheet.iter_rows(values_only=True))
        for row in rows:
            if not row:
                continue
            for cell in row:
                if cell and ("lab no" in str(cell).lower() or "lab-" in str(cell).lower()):
                    lab_info = normalize_room_number(str(cell))
                    if lab_info["room_number"] and not any(r["room_number"] == lab_info["room_number"] for r in result["rooms"]):
                        result["rooms"].append({
                            "room_number": lab_info["room_number"],
                            "room_type": "LAB",
                            "capacity": 30,
                            "has_projector": True,
                            "has_smart_board": True,
                            "building_name": "MCA Building",
                            "capabilities": ["C", "Data Structures", "Python", "AI/ML"]
                        })
